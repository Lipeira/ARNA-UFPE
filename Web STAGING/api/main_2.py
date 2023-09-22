import json
from flask import Flask, render_template, request, redirect, send_from_directory, url_for, session
from functools import wraps
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XlImage
from graphviz import Digraph
import networkx as nx
import io
import pandas as pd
from scipy.stats import norm
import matplotlib.pyplot as plt
import numpy as np


app = Flask(__name__, static_folder='static')

''
# Decorator para verificar a autenticação do usuário
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Lógica de verificação da autenticação do usuário
        if not is_user_authenticated():
            # Redirecionar o usuário para a página de login se não estiver autenticado
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def is_user_authenticated():
    return session.get('authenticated', False)

def authenticate_user():
    session['authenticated'] = True


def realizar_analise(df, Tipo):


    #CERTÍSSIMO!
    
    if Tipo == 'PERT':
        def is_edge_in_critical_path(u, v, caminho_critico):
            return (u, v) in zip(caminho_critico[:-1], caminho_critico[1:])
        
        G = nx.DiGraph()
        df['t_esperado'] = (df['t_otimista'] + 4 * df['t_provavel'] + df['t_pessimista']) / 6
        df['std'] = (df['t_pessimista'] - df['t_otimista']) / 6
        df['variance'] = ((df['t_pessimista'] - df['t_otimista']) / 6)**2

        for _, row in df.iterrows():
            G.add_node(row['Atividade'], duracao=row['t_esperado'])
            precedentes = row['precedentes']
            for precedente in precedentes:
                G.add_edge(precedente, row['Atividade'], weight=row['t_esperado'])

        caminho_critico = nx.dag_longest_path(G)

        dot = Digraph()
        dot.attr(rankdir='LR')
        dot.attr('graph', ranksep='1.5')
        dot.attr('graph', nodesep='0.75')

        for node, data in G.nodes(data=True):
            label = f"{node}"
            dot.node(node, label=label, shape='circle', ports='we')

        edge_number = 1
        edges_df = pd.DataFrame(columns=['Aresta', 'Atividade_Saida', 'Atividade_Chegada'])
        for u, v, d in G.edges(data=True):
            edge_label = f'a{edge_number}'
            atividade_saida = u
            atividade_chegada = v
            
            edges_df.loc[edge_number - 1] = [edge_label, atividade_saida, atividade_chegada]
            
            if is_edge_in_critical_path(u, v, caminho_critico):
                dot.edge(u + ":e", v + ":w", label=edge_label, color='red')
            else:
                dot.edge(u + ":e", v + ":w", label=edge_label)
            
            edge_number += 1

        dot.attr(label='PERT Network', labelloc='bottom')

        edge_labels = {f'a{i}': d['weight'] for i, (_, _, d) in enumerate(G.edges(data=True), start=1)}
        tempo_total_critico = sum([edge_labels[f'a{i}'] for i, edge in enumerate(caminho_critico, start=1)])

        i = len(caminho_critico)
        aresta_critico_list = []
        for x in range(i-1):
            atividade_saida = caminho_critico[x]
            atividade_chegada = caminho_critico[x+1]
            filtro = (edges_df['Atividade_Saida'] == atividade_saida) & (edges_df['Atividade_Chegada'] == atividade_chegada)
            valor_da_aresta = edges_df.loc[filtro, 'Aresta'].values[0]
            aresta_critico_list.append(valor_da_aresta)

        caminho_critico_string = '-->'.join(aresta_critico_list)


        static_dir = os.path.join(os.getcwd(), 'static')
        image_path = os.path.join(static_dir, 'rede_atividades')

        dot.render(image_path, view=False, format='png')
        nome_arquivo = 'df.html'

        caminho_completo = os.path.join(static_dir, nome_arquivo)

        # Salvar o DataFrame como arquivo HTML no caminho especificado
        df.to_html(caminho_completo, index=False)  


        nome_arquivo = 'edges.html'
        caminho_completo = os.path.join(static_dir, nome_arquivo)

        edges_df.to_html(caminho_completo, index=False)  

        tempo_total_critico = float(f'{tempo_total_critico:.2f}')

        return caminho_critico_string, tempo_total_critico
    
    #MANUTENÇÃO

    elif Tipo == 'CPM':
        input(df)
""" 
        def is_edge_in_critical_path(u, v, caminho_critico):
            return (u, v) in zip(caminho_critico[:-1], caminho_critico[1:])

        def calcular_es_ls(G):
            es = {}
            ls = {}

            # Calcular Early Start (ES)
            for node in nx.topological_sort(G):
                if not G.in_edges(node):
                    es[node] = 0
                else:
                    es[node] = max(es[pred] + G.edges[pred, node]['weight'] for pred in G.predecessors(node))

            # Calcular Late Start (LS)
            for node in reversed(list(nx.topological_sort(G))):
                if not G.out_edges(node):
                    ls[node] = es[node]
                else:
                    ls[node] = min(ls[succ] - G.edges[node, succ]['weight'] for succ in G.successors(node))

            return es, ls

        # Definir os dados das atividades

        # Criar o grafo direcionado
        G = nx.DiGraph()
        

        # Adicionar os nós e as arestas ao grafo
        for atividade, info in atividades.items():
            duracao = info.get('duracao', 0)
            G.add_node(atividade, duracao=duracao)
            precedentes = info.get('precedentes', [])
            for precedente in precedentes:
                G.add_edge(precedente, atividade, weight=duracao)



        # Calcular o caminho crítico
        caminho_critico = nx.dag_longest_path(G)

        # Calcular Early Start (ES) e Late Start (LS)
        es, ls = calcular_es_ls(G)

        # Calcular Early Finish (ES) e Late Finish (LS)
        ef = {node: es[node] + G.nodes[node]['duracao'] for node in G.nodes()}
        lf = {node: ls[node] + G.nodes[node]['duracao'] for node in G.nodes()}

        # Calcular a folga (float) de cada atividade
        folga = {node: ls[node] - es[node] for node in G.nodes()}


        # Calcular a posição dos nós usando o layout planar com scale=2.0
        pos = nx.planar_layout(G, scale=7.0)

        # Desenhar o grafo usando Graphviz
        dot = Digraph()
        dot.attr(rankdir='LR')  # Configurar o layout da esquerda para a direita
        dot.attr('graph', ranksep='1.5')  # Adicionar o atributo ranksep aqui (aumentar o valor para aumentar o espaço entre os níveis)
        dot.attr('graph', nodesep='0.75')  # Adicionar o atributo nodesep aqui (aumentar o valor para aumentar o espaço entre os nós no mesmo nível)

        # Adicionar os nós e as arestas ao grafo
        for node, data in G.nodes(data=True):
            label = f"{node} | Peso: {data['duracao']} | ES: {es[node]} / EF: {ef[node]} | LS: {ls[node]} / LF: {lf[node]} | Folga: {folga[node]} "
            dot.node(node, label=label, shape='record', ports='we')  # Adicionar atributo ports='we' aqui

        initial_nodes = [node for node in G.nodes() if not G.in_edges(node)]
        final_nodes = [node for node in G.nodes() if not G.out_edges(node)]

        dot.node('start', label='start', shape='rectangle')  # Adicionar o nó "start"
        dot.node('end', label='end', shape='rectangle')  # Adicionar o nó "end"

        for initial_node in initial_nodes:
            dot.edge('start', initial_node + ':w')  # Conectar "start" aos nós iniciais

        for final_node in final_nodes:
            dot.edge(final_node + ':e', 'end')  # Conectar os nós finais ao nó "end"

        label = "Tarefa|Peso|ES e EF|LS e LF|Folga"

        # Criar um subgrafo com o mesmo rank
        with dot.subgraph() as same_rank:
            same_rank.attr(rank='same')
            same_rank.node('start')  # Mover o nó "start" para este subgrafo
            # Adicionar o nó de legenda com atributo "pos" para posicioná-lo abaixo do nó "start"
            same_rank.node('legend', label=label, shape='record', pos='0,1!', pin='true') 


        for u, v, d in G.edges(data=True):
            if is_edge_in_critical_path(u, v, caminho_critico):
                dot.edge(u + ":e", v + ":w", color='red')  # Modificar u para u + ":e" e v para v + ":w" aqui
            else:
                dot.edge(u + ":e", v + ":w")  # Modificar u para u + ":e" e v para v + ":w" aqui


        dot.attr(label='CPM Network', labelloc='bottom')

        # Salvar a imagem do grafo em formato PNG
        static_dir = os.path.join(os.getcwd(), 'static')
        image_path = os.path.join(static_dir, 'rede_atividades')
        dot.render(image_path, view=False, format='png')

        generate_excel_report(atividades, caminho_critico, es, ef, ls, lf, folga)
 """



def generate_excel_report(atividades, caminho_critico, es, ef, ls, lf, folga):
    # Criar um arquivo Excel usando o Openpyxl
    wb = Workbook()
    ws = wb.active

    # Adicionar cabeçalhos
    headers = ["Atividade", "Duração", "ES", "EF", "LS", "LF", "Folga", "Caminho Crítico"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Preencher dados das atividades
    row = 2
    for atividade, info in atividades.items():
        duracao = info.get("duracao", "")
        es_value = es.get(atividade, "")
        ef_value = ef.get(atividade, "")
        ls_value = ls.get(atividade, "")
        lf_value = lf.get(atividade, "")
        folga_value = folga.get(atividade, "")
        caminho_critico_value = "Sim" if atividade in caminho_critico else "Não"
        ws[f"A{row}"] = atividade
        ws[f"B{row}"] = duracao
        ws[f"C{row}"] = es_value
        ws[f"D{row}"] = ef_value
        ws[f"E{row}"] = ls_value
        ws[f"F{row}"] = lf_value
        ws[f"G{row}"] = folga_value
        ws[f"H{row}"] = caminho_critico_value
        row += 1

    # Autoajuste das larguras das colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

     # Salvar o arquivo Excel
    wb.save("static/report.xlsx")

app.secret_key = 'ALDmofmognmg15448d8'
# ------------ ROTAS -----------------

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Lógica de autenticação do usuário
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        if email == 'admin' and senha == '123456':
            # Autenticar o usuário
            authenticate_user() 
            return redirect(url_for('home'))
        else:
            mensagem_erro = 'Usuário ou senha incorretos'
            return render_template('login.html', erro=mensagem_erro)

    return render_template('login.html')


@app.route('/home')
@login_required
def home():
    return render_template('home.html')

@app.route('/PERT')
@login_required
def PERT():
    return render_template('homePERT.html')


@app.route('/analyze', methods=['POST'])
@login_required
def analyze():
    if request.method == 'POST':
            try:
                csv_file = request.files['csv_file']  # Obter o arquivo CSV do formulário
            except:
                csv_file = False

            try:
                json_file = request.files['json_file'] # Obter o arquivo JSON do formulário
            except:
                json_file = False



            if json_file:  # Se um arquivo CSV foi enviado
                # Carrega o arquivo JSON em um DataFrame
                
                try:
                    caminho_arquivo = 'atividades.json'
                    data_frame = pd.read_json(caminho_arquivo)
                    data_frame = data_frame.transpose()
                except:
                    erro = "Ocorreu um erro!"
                    return render_template("homePERT.html", erro=erro)

                # Transformar o índice em uma coluna
                data_frame = data_frame.reset_index()

                # Renomear a coluna do índice para 'atividades'
                data_frame.rename(columns={'index': 'Atividade'}, inplace=True)
                atividades_json = data_frame
                
            elif csv_file:
                try:
                    caminho_arquivo = 'atividades.csv'
                    data_frame = pd.read_csv(caminho_arquivo)
                except:
                    erro = "Ocorreu um erro!"
                    return render_template("homePERT.html", erro=erro)

                atividades_json = data_frame

        
            else:
                atividades = request.form.get("atividades")
                # Verificar se o campo de atividades está vazio
                if not atividades:
                    erro = "O campo de atividades não pode estar vazio."
                    return render_template("home.html", erro=erro)
                
                atividades_json = json.loads(atividades)
                atividades_json = pd.DataFrame.from_dict(atividades_json, orient='index')
                # Exibir o DataFrame resultante
            
                
            caminho_critico, tempo_total = realizar_analise(atividades_json, 'CPM')

            session['caminho_critico'] = caminho_critico
            session['tempo_total'] = tempo_total

            return redirect(url_for('result'))
        

@app.route('/analyzePERT', methods=['POST'])
@login_required
def analyzePERT():
    if request.method == 'POST':
    
        try:
            csv_file = request.files['csv_file']  # Obter o arquivo CSV do formulário
        except:
            csv_file = False

        try:
            json_file = request.files['json_file'] # Obter o arquivo JSON do formulário
        except:
            json_file = False



        if json_file:  # Se um arquivo CSV foi enviado
            # Carrega o arquivo JSON em um DataFrame
            
            try:
                caminho_arquivo = 'atividades.json'
                data_frame = pd.read_json(caminho_arquivo)
                data_frame = data_frame.transpose()
            except:
                erro = "Ocorreu um erro!"
                return render_template("homePERT.html", erro=erro)

            # Transformar o índice em uma coluna
            data_frame = data_frame.reset_index()

            # Renomear a coluna do índice para 'atividades'
            data_frame.rename(columns={'index': 'Atividade'}, inplace=True)
            atividades_json = data_frame
            
        elif csv_file:
            try:
                caminho_arquivo = 'atividades.csv'
                data_frame = pd.read_csv(caminho_arquivo)
            except:
                erro = "Ocorreu um erro!"
                return render_template("homePERT.html", erro=erro)

            atividades_json = data_frame

    
        else:
            atividades = request.form.get("atividades")
            # Verificar se o campo de atividades está vazio
            if not atividades:
                erro = "O campo de atividades não pode estar vazio."
                return render_template("home.html", erro=erro)
            
            atividades_json = json.loads(atividades)
            atividades_json = pd.DataFrame.from_dict(atividades_json, orient='index')
            # Exibir o DataFrame resultante
        
            
        caminho_critico, tempo_total = realizar_analise(atividades_json, 'PERT')

        session['caminho_critico'] = caminho_critico
        session['tempo_total'] = tempo_total

        return render_template('resultPERT.html', caminho_critico=caminho_critico, tempo_total=tempo_total)
    


@app.route('/result')
@login_required
def result():
    return render_template('result.html')


@app.route('/resultPERT')
@login_required
def resultPERT():
    caminho_critico = session.get('caminho_critico')
    tempo_total = session.get('tempo_total')
    edges_df = session.get('edges_df')

    return render_template('resultPERT.html')

@app.route('/download_png')
@login_required
def download_png():
    # Baixar o arquivo PNG do Amazon S3
    """ s3.download_file(BUCKET_NAME, 'rede_atividades.png', 'static/rede_atividades.png') """
    return send_from_directory('static', 'rede_atividades.png', as_attachment=True)

@app.route("/download_pdf")
@login_required
def download_pdf():
    # Baixar o arquivo PDF do Amazon S3
    """ s3.download_file(BUCKET_NAME, 'report.pdf', 'static/report.pdf') """
    return send_from_directory('static', 'report.pdf', as_attachment=True)

@app.route("/download_xls")
@login_required
def download_xls():
    # Baixar o arquivo XLSX do Amazon S3
    """ s3.download_file(BUCKET_NAME, 'report.xlsx', 'static/report.xlsx') """
    return send_from_directory('static', 'report.xlsx', as_attachment=True)


@app.route('/help')
@login_required
def help():
    return render_template('help.html')

@app.route('/contact')
@login_required
def contact():
    return render_template('contact.html')

@app.route('/gauss', methods=['POST'])
@login_required
def gauss():
    t_programado = float(request.form.get("t_programado"))
    tempo_total = session.get('tempo_total')
    atividades_json = session.get('atividades_json')

    soma_var = 0
    for act in atividades_json:
        soma_var += atividades_json[act]['variance']

    z = (t_programado - tempo_total) / (soma_var)**(1/2)
    z = f"{z:.1f}"
    z = float(z)

    mi = tempo_total
    pct_projeto_programado = 50/100

    area = norm.cdf(z)
    print(f"Área: {area:.3f}")

    # Cria um conjunto de valores x no intervalo [tempo_total - 4, tempo_total + 4] para o tempo do projeto
    x = np.linspace(tempo_total - 5, tempo_total + 5, 1000)

    # Calcula a probabilidade acumulada até o valor Z
    area = norm.cdf(z)

    # Cria uma figura e um eixo
    fig, ax = plt.subplots()

    # Plota a curva gaussiana
    ax.plot(x, norm.pdf(x, tempo_total), label='Distribuição Normal')

    # Preenche a área sob a curva até o valor de Z
    ax.fill_between(x, 0, norm.pdf(x, tempo_total), where=(x <= z + tempo_total), alpha=0.3, label=f'Probabilidade = {area:.5f}')

    # Adiciona uma linha vermelha no valor médio (símbolo μ)
    ax.axvline(x=tempo_total, color='red', linestyle='dashed', label=f'Valor Médio (μ = {tempo_total:.2f})')

    # Configurações do gráfico
    ax.set_title('Distribuição Gaussiana com Valor Médio')
    ax.set_xlabel('Tempo do Projeto')
    ax.set_ylabel('Densidade de Probabilidade')
    ax.legend()

    static_dir = os.path.join(os.getcwd(), 'static')
    plt.savefig(os.path.join(static_dir, 'grafico_gaussiano.png'))
    return send_from_directory('static', 'grafico_gaussiano.png', as_attachment=True)


if __name__ == '__main__':
    app.run()